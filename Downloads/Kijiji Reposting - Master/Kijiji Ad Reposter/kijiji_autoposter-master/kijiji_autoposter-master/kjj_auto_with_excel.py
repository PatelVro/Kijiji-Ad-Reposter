import os
import sys
import time
import datetime
import random
from selenium import webdriver
from selenium.webdriver.firefox.options import Options as FirefoxOptions
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.common.keys import Keys
from selenium.common.exceptions import ElementClickInterceptedException
from selenium.webdriver.chrome.options import Options as ChromeOptions
from openpyxl import load_workbook

class kijiji():
    
    def __init__(self):
        self.current_dir = os.getcwd()
        self.headless = ChromeOptions()
        self.headless.add_argument("--headless") # Set headless mode within the FirefoxOptions object

    @staticmethod
    def is_page_fully_loaded(driver):
        return driver.execute_script("return document.readyState") == "complete"

    def access_kijiji(self):
        # Set the timeout for the entire script execution to 60 seconds
        
        self.kjj = webdriver.Chrome(options=self.headless)

        while True:
            try:
                wait = WebDriverWait(self.kjj, 20)
                self.next_url('https://www.kijiji.ca/')
                wait = WebDriverWait(self.kjj, 20)
                sign_in_link = wait.until(EC.element_to_be_clickable((By.LINK_TEXT, "Sign In")))
                sign_in_link.click()
                print("Successfully found and clicked on the 'Sign In' link.")
                # Proceed with the rest of your code for logging in
                self.login()
                break  # Exit the loop if successful
            except TimeoutException:
                print("Timeout occurred. Unable to find or click on the 'Sign In' link. Retrying...")
                self.kjj.refresh()
            except ElementClickInterceptedException:
                print("ElementClickInterceptedException: Unable to click on the 'Sign In' link due to an intercepted click. Retrying...")
                self.kjj.refresh()
            except Exception as e:
                print(f"An unexpected error occurred: {str(e)} Retrying...")
                self.kjj.refresh()


    def next_url(self,new_url):
        current = self.kjj.current_url
        self.kjj.get(new_url)
        wait = time.time() + 10 # add 10 second time out
        while current == self.kjj.current_url and time.time() < wait:
            time.sleep(1)
            
    def next_click(self,e_to_click):
        current = self.kjj.current_url
        e_to_click.click()
        wait = time.time() + 10 # add 10 second time out
        while current == self.kjj.current_url and time.time() < wait:
            time.sleep(1)
            
    def post_ad(self, ad_data):
        self.current_ad_title = ad_data['Title']
        self.current_categories = ad_data['Category']
        self.current_ad_price = ad_data['Price']
        self.description = ad_data['Description']
        self.condition = ad_data['Condition']
        self.phonebrand = ad_data['PhoneBrand']
        self.phonebrand_carrier = ad_data['PhoneBrandCarrier']
        self.current_folderName = ad_data['Images_FolderName']
        self.Phone = ad_data['Phone']
        self.tags = ad_data['Tags']
        self.size = ad_data['Size']
        self.type = ad_data['Type']
        self.tablet_brand = ad_data['Tablet Brand']
        self.laptop_Screen_Size = ad_data['laptop Screen Size']

        self.kjj.refresh()
        time.sleep(5)
        wait = WebDriverWait(self.kjj, 60)
        post_ad_link = wait.until(EC.element_to_be_clickable((By.LINK_TEXT, "Post ad")))
        post_ad_link.click()
        
        WebDriverWait(self.kjj, 60).until(EC.presence_of_element_located((By.ID, "AdTitleForm")))
        title = self.kjj.find_element(By.ID,'AdTitleForm')
        title.send_keys(self.current_ad_title)
        
        wait = WebDriverWait(self.kjj, 60)
        next_button_link = wait.until(EC.element_to_be_clickable((By.TAG_NAME, "button")))
        next_button_link.click()
        
        elements = self.kjj.find_elements(By.XPATH, '//*[contains(@class, "suggestedCategory")]')
        if elements:
            elements[0].click()
            print('Suggested Category Clicked')
        else:
            categories = self.current_categories.split(',')
            for category in categories:
                try:
                    xpath_expression = f'//button[contains(., "{category.strip()}")]'
                    categories_link = WebDriverWait(self.kjj, 20).until(EC.element_to_be_clickable((By.XPATH, xpath_expression)))
                    categories_link.click()
                    print(category)
                except TimeoutException:
                    print(f"Timed out while waiting for {category}, continuing with the next category...")
                    continue
        
        print('Category selected')
        time.sleep(5)
        
        try:
            label_element = WebDriverWait(self.kjj, 10).until(
                EC.element_to_be_clickable((By.CSS_SELECTOR, 'label[for="forsaleby_s-delr"]'))
            )
            label_element.click()
        except TimeoutException:
            print("Element for Business Radio Option Not Found!")
            
        try:    
            label_element = WebDriverWait(self.kjj, 10).until(
                EC.element_to_be_clickable((By.CSS_SELECTOR, 'label[for="fulfillment_s-delivery"]'))
            )
            label_element.click()
        except TimeoutException:
            print("Element for Delivery Checkbox Not Found!")
        
        try:
            label_element = WebDriverWait(self.kjj, 10).until(
                EC.element_to_be_clickable((By.CSS_SELECTOR, 'label[for="fulfillment_s-shipping"]'))
            )
            label_element.click()
        except TimeoutException:
            print("Element for Shipping Checkbox Not Found!")
        
        try:
            label_element = WebDriverWait(self.kjj, 10).until(
                EC.element_to_be_clickable((By.CSS_SELECTOR, 'label[for="fulfillment_s-curbside"]'))
            )
            label_element.click()
        except TimeoutException:
            print("Element for Curbside Checkbox Not Found!")
          
        try:    
            label_element = WebDriverWait(self.kjj, 10).until(
                EC.element_to_be_clickable((By.CSS_SELECTOR, 'label[for="payment_s-cashless"]'))
            )
            label_element.click()
        except TimeoutException:
            print("Element for Cashless payment Checkbox Not Found!")
        
        try:
            label_element = WebDriverWait(self.kjj, 10).until(
                EC.element_to_be_clickable((By.CSS_SELECTOR, 'label[for="payment_s-cashaccepted"]'))
            )
            label_element.click()
        except TimeoutException:
            print("Element for Cash payment Checkbox Not Found!")
        
        try:
            # Locate the dropdown element
            condition_dropdown = self.kjj.find_element(By.ID, 'condition_s')
            # Create a Select object
            select = Select(condition_dropdown)

            # Select the option by its value
            select.select_by_value(self.condition)
        except NoSuchElementException:
            print("Element for Condition Dropdown Not Found!")

        try:
            # Locate the dropdown element
            phonebrand_dropdown = self.kjj.find_element(By.ID, 'phonebrand_s')

            # Create a Select object
            select = Select(phonebrand_dropdown)

            # Select the option by its value
            select.select_by_value(self.phonebrand)
        except NoSuchElementException:
            print("Element for Phone Brand Dropdown Not Found!")
                        
        try:
                # Locate the dropdown element
            phone_Carrier_dropdown = self.kjj.find_element(By.ID, 'phonecarrier_s')

            # Create a Select object
            select = Select(phone_Carrier_dropdown)

            # Select the option by its value
            select.select_by_value(self.phonebrand_carrier)
        except NoSuchElementException:
            print("Element for Phone Brand Carrier Dropdown Not Found!")

        try:
            # Locate the dropdown element
            phonebrand_dropdown = self.kjj.find_element(By.ID, 'phonebrand_s')

            # Create a Select object
            select = Select(phonebrand_dropdown)

            # Select the option by its value
            select.select_by_value(self.phonebrand)
        except NoSuchElementException:
            print("Element for Phone Brand Dropdown Not Found!")
                        
        try:
                # Locate the dropdown element
            phone_Carrier_dropdown = self.kjj.find_element(By.ID, 'phonecarrier_s')

            # Create a Select object
            select = Select(phone_Carrier_dropdown)

            # Select the option by its value
            select.select_by_value(self.phonebrand_carrier)
        except NoSuchElementException:
            print("Element for Phone Brand Carrier Dropdown Not Found!")
        
        # Locate the dropdown element
        try:
            condition_dropdown = self.kjj.find_element(By.ID, 'condition_s')
            select = Select(condition_dropdown)
            select.select_by_value(self.condition)
        except NoSuchElementException:
            print("Condition dropdown not found. Skipping...")

        # Locate the dropdown element
        try:
            Screen_Size_dropdown = self.kjj.find_element(By.ID, 'size_s')
            select = Select(Screen_Size_dropdown)
            select.select_by_value(self.size)
        except NoSuchElementException:
            print("Size dropdown not found. Skipping...")

        # Locate the dropdown element
        try:
            TV_Type_dropdown = self.kjj.find_element(By.ID, 'type_s')
            select = Select(TV_Type_dropdown)
            select.select_by_value(self.type)
        except NoSuchElementException:
            print("TV Type dropdown not found. Skipping...")

        try:
            Tablet_Brand_Dropdown = self.kjj.find_element(By.ID, 'tabletbrand_s')
            select = Select(Tablet_Brand_Dropdown)
            select.select_by_value(self.tablet_brand)
        except NoSuchElementException:
            print("Tablet Brand dropdown not found. Skipping...")   

                # Locate the dropdown element
        try:
            Laptop_Screen_Size_dropdown = self.kjj.find_element(By.ID, 'laptopscreensize_s')
            select = Select(Laptop_Screen_Size_dropdown)
            select.select_by_value(self.laptop_Screen_Size)
        except NoSuchElementException:
            print("Size dropdown not found. Skipping...")         

        # Choose and enter ad price if present
        try:
            WebDriverWait(self.kjj, 60).until(EC.presence_of_element_located((By.ID, "PriceAmount")))
            price = self.kjj.find_element(By.ID,'PriceAmount')
            price.send_keys(self.current_ad_price)
        except NoSuchElementException:
            print("Price input field not found. Skipping...")

        # Enter description if present
        try:
            desc = self.kjj.find_element(By.TAG_NAME,'textarea')
            desc.send_keys(self.description)
        except NoSuchElementException:
            print("Description input field not found. Skipping...")

        # Upload images if present
        try:
            script_dir = os.path.dirname(os.path.abspath(__file__))

            folder_path = os.path.normpath(os.path.join(script_dir, '../../../Kijiji Ad Photos/Kijiji Ad Photos/', self.current_folderName  ))
            # folder_path = os.path.abspath(os.path.join('C:/Users/user/Documents/Kijiji Ad Photos', self.current_folderName))
            image_list = os.listdir(folder_path)
            upload = self.kjj.find_element(By.CLASS_NAME, 'imageUploadButtonWrapper')
            upload = upload.find_element(By.TAG_NAME, 'input')

            for pic in image_list:
                absolute_path = os.path.join(folder_path, pic)  # Construct absolute path to each image
                print("Absolute path:", absolute_path)
                upload.send_keys(absolute_path)

            print("Photos Uploaded")
        except NoSuchElementException:
            print("Image upload button not found. Skipping...")

        # Wait for the element with ID "PhoneNumber" to be visible
        try:
            WebDriverWait(self.kjj, 10).until(EC.visibility_of_element_located((By.ID, "PhoneNumber")))
            phone = self.kjj.find_element(By.ID,'PhoneNumber')
            # Once the element is visible, send keys to it
            numbers_list = list(str(self.Phone))
            for number in numbers_list:
                phone.send_keys(int(number))
            
    
        except NoSuchElementException:
            print("Phone number input field not found. Skipping...")
        
        try:
            tags_list = self.tags.split(',')
            tags = self.kjj.find_element(By.ID,'pstad-tagsInput')
        
            for tag in tags_list:
                tags.send_keys(tag)
                tags.send_keys(Keys.RETURN)  # Press ENTER key to submit the tag
        except NoSuchElementException:
            print("Tags input field not found. Skipping...")
        
        time.sleep(20)
        post_ad_link = wait.until(EC.element_to_be_clickable((By.XPATH, '//button[text()="Post Your Ad"]')))
        post_ad_link.click()
        time.sleep(20)        
        
        print(
            'Ad posted successfully.',
            '\nCurrent title is {}.'.format(self.current_ad_title),
            '\nCurrent price is {}.'.format(self.current_ad_price)
        )

    def login(self):
        time.sleep(5)
        while True:
            try:
                WebDriverWait(self.kjj, 60).until(EC.presence_of_element_located((By.ID, "username")))
                user = self.kjj.find_element(By.ID,'username')
                user.send_keys("hillpatel1357@gmail.com")
                
                WebDriverWait(self.kjj, 60).until(EC.presence_of_element_located((By.ID, "password")))
                password = self.kjj.find_element(By.ID,'password')
                password.send_keys("pbR.fSb^G^Sku5g")
                                
                WebDriverWait(self.kjj, 60).until(EC.presence_of_element_located((By.ID, "login-submit")))
                self.next_click(self.kjj.find_element(By.ID, "login-submit"))
                break
            except TimeoutException:
                print("TimeoutException: One or more elements did not load within 60 seconds.")
            # Navigate back using browser navigation commands
                self.kjj.back()
                print("Navigated back to the previous page.")
            except Exception as e:
                print(f"An unexpected error occurred: {str(e)}")
    
    def delete_ad(self,ad_data):
        wait = WebDriverWait(self.kjj, 60)
        self.current_ad_title = ad_data['Title']
        self.next_url('https://www.kijiji.ca/m-my-ads/active/1')
        time.sleep(5)   
        div_elements = []
        try:
            tbody_element = self.kjj.find_element(By.TAG_NAME, 'tbody')
            div_elements = tbody_element.find_elements(By.XPATH,'.//div[not(@class)]')            
        except NoSuchElementException:
            print("Element not found.")
            # Handle the situation here, such as returning from the method or raising an error

        found_ad_to_delete = False
        if not div_elements:
            print("No div elements found.")
        else:        
            for div_element in div_elements:
                links = div_element.find_elements(By.XPATH, './/a[@href]')
                for link in links:
                    if self.current_ad_title in link.get_attribute("innerHTML"):
                        try:
                            delete_button = div_element.find_element(By.XPATH, './/span[text()="Delete"]/ancestor::button')
                            delete_button.click()
                            found_ad_to_delete = True
                            break
                        except NoSuchElementException:
                            print("Delete button not found for the ad.")
                if found_ad_to_delete:
                    break

        if found_ad_to_delete:
            try:
                button = WebDriverWait(self.kjj, 10).until(EC.element_to_be_clickable((By.XPATH, '//button[text()="Prefer not to say"]')))
                button.click()
                time.sleep(2)
                button = WebDriverWait(self.kjj, 10).until(EC.element_to_be_clickable((By.XPATH, '//button[text()="Delete My Ad"]')))
                button.click()
                time.sleep(2)
                print('Ad deleted.')
                button = WebDriverWait(self.kjj, 10).until(EC.element_to_be_clickable((By.ID, 'modalCloseButton')))
                button.click()
                time.sleep(2)
                self.kjj.refresh()
            except TimeoutException:
                print('Timeout occurred while deleting the ad.')
        else:
            print('Ad with title "{}" not found for deletion.'.format(self.current_ad_title))
            

def main():
    # for i in range(1, 721):  # Start from 1 and end at 720
    #     print(f"Iteration {i}")
    #     time.sleep(60)

    browser = kijiji()
    browser.access_kijiji()

    try:
        # Load data from Excel file
        print("About to Start")
        wb = load_workbook('ads_data_new.xlsx')
        print("Ads_Data Loaded")
        sheet = wb.active

        # Start from the second row (assuming the first row is headers)
        for row in sheet.iter_rows(min_row=2, values_only=True):
            ad_data = {
                'Title': row[0],
                'Category': row[1],
                'Price': row[2],
                'Description': row[3],
                'Condition': row[4],
                'PhoneBrand': row[5],
                'PhoneBrandCarrier': row[6],
                'Images_FolderName': row[7],
                'Phone': row[8],
                'Tags': row[9], 
                'Size': row[10],
                'Type': row[11],
                'Tablet Brand': row[12],
                'laptop Screen Size': row[13]
            }
            browser.delete_ad(ad_data)
            time.sleep(10)
            browser.post_ad(ad_data)
            time.sleep(10)  # Sleep between posting ads
    except Exception as e:
        print(f"Error occurred: {e}")

if __name__ == '__main__':
    main()
