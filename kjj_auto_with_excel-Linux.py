import os
import sys
import time
import datetime
import random
import pymysql
import json
from selenium import webdriver
from selenium.webdriver.firefox.options import Options as FirefoxOptions
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.common.keys import Keys
from selenium.common.exceptions import ElementClickInterceptedException
from selenium.webdriver.chrome.options import Options as ChromeOptions
from openpyxl import load_workbook
from openpyxl import Workbook
from selenium.common.exceptions import StaleElementReferenceException
from db_config import config

class kijiji():
    
    def __init__(self):
        self.current_dir = os.getcwd()
        self.headless = ChromeOptions()
        #self.headless.add_argument("--headless") # Set headless mode within the FirefoxOptions object
        self.headless.add_argument("--disable-blink-features=AutomationControlled")
        self.headless.add_argument(r"user-data-dir=C:\Users\Administrator\AppData\Local\Google\Chrome\User Data") #Path to your chrome profile
        self.headless.add_argument(r'--profile-directory=Default')   
        self.headless.add_argument("--lang=en-US")
        self.db_connection = None     

    @staticmethod
    def is_page_fully_loaded(driver):
        return driver.execute_script("return document.readyState") == "complete"

    def connect_db(self,credentials):
        try:
            self.email = credentials['username']
            # Establish the database connection
            connection = pymysql.connect(
                host=config['host'],
                port=config['port'],
                user=config['user'],
                password=config['password'],
                database=config['database']
            )

            with connection.cursor() as cursor:
            
                # Execute a simple query
                cursor.execute("SELECT DATABASE();")
                result = cursor.fetchone()
                print("You're connected to database: ", result[0])
                
                # Execute a test query
                cursor.execute(f"SELECT * FROM `{self.email}`")
                
                # Fetch column names
                column_names = [desc[0] for desc in cursor.description]
                
                # Fetch row data
                rows = cursor.fetchall()
                
                print("Data from `{self.email}` table:")
                
                # Convert rows to a list of dictionaries
                data = [dict(zip(column_names, row)) for row in rows]
                
                json_data = json.dumps(data, indent=4)
                print(json_data)
                return json_data

                # Closing the connection
            # connection.close()
            # print("MariaDB connection is closed")
        except pymysql.MySQLError as e:
            print(f"Error while connecting to MariaDB: {e}")

    # def fetch_data(self):
    #     try:
    #         with self.db_connection.cursor() as cursor:
    #             sql = "SELECT * FROM `cbluff170@gmail.com`"
    #             cursor.execute(sql)
    #             result = cursor.fetchall()
    #             print(f"Fetched {len(result)} records from the database.")
    #             return result
    #     except Exception as e:
    #         print(f"Failed to fetch data: {str(e)}")
    #         return []

    def access_kijiji(self,credentials):
        # Set the timeout for the entire script execution to 60 seconds
        # Set preferences to allow all cookies


        # Add argument to disable automation control detection
        self.kjj = webdriver.Chrome(options=self.headless)

        while True:
            try:
                # Check if the current WebDriver is Chrome
                
                wait = WebDriverWait(self.kjj, 20)
                self.next_url('https://www.kijiji.ca/?siteLocale=en_CA')
                wait = WebDriverWait(self.kjj, 20)

                # try:
                #     # Wait up to 10 seconds for the popup to appear
                #     cookie_banner = WebDriverWait(self.kjj, 10).until(
                #         EC.presence_of_element_located((By.XPATH, "//*[@id='MainContainer']/div[1]/div/div[2]/div[2]/button"))
                #     )
                #     cookie_banner.get_attribute("outerHTML")
                    
                #     print(cookie_banner)
                #     # Find the accept button and click it
                #     # close_button = cookie_banner.find_element_by_xpath("//button[contains(text(), 'Close')]")
                #     cookie_banner.click()
                    
                #     print("Privacy policy or cookies acceptance bypassed successfully.")
                
                # except Exception as e:
                #     print(f"An unexpected error occurred: {str(e)}")
                
                # sign_in_link = wait.until(EC.element_to_be_clickable((By.LINK_TEXT, "Sign In")))
                # sign_in_link.click()
                # print("Successfully found and clicked on the 'Sign In' link.")
                # # Proceed with the rest of your code for logging in
                # self.login(credentials)
                break  # Exit the loop if successful            
            except TimeoutException:
                print("Timeout occurred. Unable to find or click on the 'Sign In' link. Retrying...")
                self.kjj.refresh()
            except ElementClickInterceptedException:
                print("ElementClickInterceptedException: Unable to click on the 'Sign In' link due to an intercepted click. Retrying...")
                self.kjj.refresh()
            except Exception as e:
                print(f"An unexpected error occurred: {str(e)} Retrying...")
                self.kjj.refresh()


    def next_url(self,new_url):
        current = self.kjj.current_url
        self.kjj.get(new_url)
        wait = time.time() + 10 # add 10 second time out
        while current == self.kjj.current_url and time.time() < wait:
            time.sleep(1)
            
    def next_click(self,e_to_click):
        current = self.kjj.current_url
        e_to_click.click()
        wait = time.time() + 10 # add 10 second time out
        while current == self.kjj.current_url and time.time() < wait:
            time.sleep(1)
          
    def copy_sheet(self, source_file, source_sheet_index, destination_file):
        try:
            # Ensure source_sheet_index is an integer
            source_sheet_index = int(source_sheet_index)
            
            # Load the source Excel file
            source_wb = load_workbook(source_file)
            # Get the source sheet by index
            source_sheet = source_wb.worksheets[source_sheet_index - 1]  # Index is 0-based
            
            # Check if the destination file exists
            if not os.path.exists(destination_file):
                # If it doesn't exist, create a new workbook
                dest_wb = Workbook()
                dest_wb.save(destination_file)
            
            # Load the destination workbook
            dest_wb = load_workbook(destination_file)
            # Get the active sheet (which is the first sheet by default)
            dest_sheet = dest_wb.active
            
            # Clear contents of the destination sheet
            dest_sheet.delete_rows(1, dest_sheet.max_row)  # Delete all rows
            
            # Copy data from source sheet to destination sheet
            for row in source_sheet.iter_rows(values_only=True):
                dest_sheet.append(row)
            
            # Save the destination workbook
            dest_wb.save(destination_file)
            
            print(f"Sheet at index {source_sheet_index} copied from '{source_file}' to '{destination_file}' successfully.")
        except Exception as e:
            print(f"An error occurred: {str(e)}")




          
            
    def post_ad(self, ad_data):
        try:
            self.current_ad_title = ad_data['Title'].strip()
            self.current_categories = ad_data['Category']
            self.current_ad_price = ad_data['Price']
            self.description = ad_data['Description']
            self.condition = ad_data['Condition']
            self.phonebrand = ad_data['PhoneBrand']
            self.phonebrand_carrier = ad_data['PhoneBrandCarrier']
            self.current_folderName = ad_data['Images_FolderName'].strip()
            self.Phone = ad_data['Phone']
            self.tags = ad_data['Tags']
            self.size = ad_data['Size']
            self.type = ad_data['Type']
            self.tablet_brand = ad_data['Tablet Brand']
            self.laptop_Screen_Size = ad_data['laptop Screen Size']

            self.kjj.refresh()
            time.sleep(5)
            wait = WebDriverWait(self.kjj, 60)
            post_ad_link = wait.until(EC.element_to_be_clickable((By.LINK_TEXT, "Post ad")))
            post_ad_link.click()
            
            WebDriverWait(self.kjj, 60).until(EC.presence_of_element_located((By.ID, "AdTitleForm")))
            title = self.kjj.find_element(By.ID,'AdTitleForm')
            title.send_keys(self.current_ad_title.strip())
            
            wait = WebDriverWait(self.kjj, 60)
            next_button_link = wait.until(EC.element_to_be_clickable((By.TAG_NAME, "button")))
            next_button_link.click()

            categories = self.current_categories.split(';')

            time.sleep(5)

            for category in categories:
                print(category + " Iteration.....")
                # Find the container element
                container = self.kjj.find_element(By.CLASS_NAME, "allCategoriesContainer-1711250309")

                # Find all buttons within the container
                buttons = container.find_elements(By.XPATH, ".//button")
                
                for button in buttons:
                    try:
                        category_name = button.find_element(By.TAG_NAME, "span").text.strip()
                        if category_name == category.strip():
                            # Click the button
                            print("Match Found")
                            button.click()
                            time.sleep(2)
                            break
                    except NoSuchElementException:
                        continue

            
            print('Category selected')
            time.sleep(5)
            
            try:
                label_element = WebDriverWait(self.kjj, 10).until(
                    EC.element_to_be_clickable((By.CSS_SELECTOR, 'label[for="forsaleby_s-delr"]'))
                )
                label_element.click()
            except TimeoutException:
                print("Element for Business Radio Option Not Found!")
            except Exception as e:
                print(f"An unexpected error occurred: {str(e)}")
                
            try:    
                label_element = WebDriverWait(self.kjj, 10).until(
                    EC.element_to_be_clickable((By.CSS_SELECTOR, 'label[for="fulfillment_s-delivery"]'))
                )
                label_element.click()
            except TimeoutException:
                print("Element for Delivery Checkbox Not Found!")
            except Exception as e:
                print(f"An unexpected error occurred: {str(e)}")
            
            try:
                label_element = WebDriverWait(self.kjj, 10).until(
                    EC.element_to_be_clickable((By.CSS_SELECTOR, 'label[for="fulfillment_s-shipping"]'))
                )
                label_element.click()
            except TimeoutException:
                print("Element for Shipping Checkbox Not Found!")
            except Exception as e:
                print(f"An unexpected error occurred: {str(e)}")
            
            try:
                label_element = WebDriverWait(self.kjj, 10).until(
                    EC.element_to_be_clickable((By.CSS_SELECTOR, 'label[for="fulfillment_s-curbside"]'))
                )
                label_element.click()
            except TimeoutException:
                print("Element for Curbside Checkbox Not Found!")
            except Exception as e:
                print(f"An unexpected error occurred: {str(e)}")
            
            try:    
                label_element = WebDriverWait(self.kjj, 10).until(
                    EC.element_to_be_clickable((By.CSS_SELECTOR, 'label[for="payment_s-cashless"]'))
                )
                label_element.click()
            except TimeoutException:
                print("Element for Cashless payment Checkbox Not Found!")
            except Exception as e:
                print(f"An unexpected error occurred: {str(e)}")
            
            try:
                label_element = WebDriverWait(self.kjj, 10).until(
                    EC.element_to_be_clickable((By.CSS_SELECTOR, 'label[for="payment_s-cashaccepted"]'))
                )
                label_element.click()
            except TimeoutException:
                print("Element for Cash payment Checkbox Not Found!")
            except Exception as e:
                print(f"An unexpected error occurred: {str(e)}")
            
            try:
                # Locate the dropdown element
                condition_dropdown = self.kjj.find_element(By.ID, 'condition_s')
                # Create a Select object
                select = Select(condition_dropdown)

                # Select the option by its value
                if self.condition:
                    select.select_by_value(self.condition)
            except NoSuchElementException:
                print("Element for Condition Dropdown Not Found!")
            except Exception as e:
                print(f"An unexpected error occurred: {str(e)}")

            try:
                # Locate the dropdown element
                phonebrand_dropdown = self.kjj.find_element(By.ID, 'phonebrand_s')

                # Create a Select object
                select = Select(phonebrand_dropdown)

                # Select the option by its value
                if self.phonebrand:
                    select.select_by_value(self.phonebrand)
            except NoSuchElementException:
                print("Element for Phone Brand Dropdown Not Found!")
            except Exception as e:
                print(f"An unexpected error occurred: {str(e)}")
                            
            try:
                    # Locate the dropdown element
                phone_Carrier_dropdown = self.kjj.find_element(By.ID, 'phonecarrier_s')

                # Create a Select object
                select = Select(phone_Carrier_dropdown)

                # Select the option by its value
                if self.phonebrand_carrier:
                    select.select_by_value(self.phonebrand_carrier)
            except NoSuchElementException:
                print("Element for Phone Brand Carrier Dropdown Not Found!")
            except Exception as e:
                print(f"An unexpected error occurred: {str(e)}")

            # Locate the dropdown element
            try:
                condition_dropdown = self.kjj.find_element(By.ID, 'condition_s')
                select = Select(condition_dropdown)
                if self.condition:
                    select.select_by_value(self.condition)
            except NoSuchElementException:
                print("Condition dropdown not found. Skipping...")
            except Exception as e:
                print(f"An unexpected error occurred: {str(e)}")

            # Locate the dropdown element
            try:
                Screen_Size_dropdown = self.kjj.find_element(By.ID, 'size_s')
                select = Select(Screen_Size_dropdown)

                if self.size:
                    select.select_by_value(self.size)
            except NoSuchElementException:
                print("Size dropdown not found. Skipping...")
            except Exception as e:
                print(f"An unexpected error occurred: {str(e)}")

            # Locate the dropdown element
            try:
                TV_Type_dropdown = self.kjj.find_element(By.ID, 'type_s')
                select = Select(TV_Type_dropdown)

                if self.type:
                    select.select_by_value(self.type)
            except NoSuchElementException:
                print("TV Type dropdown not found. Skipping...")
            except Exception as e:
                print(f"An unexpected error occurred: {str(e)}")

            try:
                Tablet_Brand_Dropdown = self.kjj.find_element(By.ID, 'tabletbrand_s')
                select = Select(Tablet_Brand_Dropdown)
                if self.tablet_brand:
                    select.select_by_value(self.tablet_brand)
            except NoSuchElementException:
                print("Tablet Brand dropdown not found. Skipping...")  
            except Exception as e:
                print(f"An unexpected error occurred: {str(e)}") 

                    # Locate the dropdown element
            try:
                Laptop_Screen_Size_dropdown = self.kjj.find_element(By.ID, 'laptopscreensize_s')
                select = Select(Laptop_Screen_Size_dropdown)
                if self.laptop_Screen_Size:
                    select.select_by_value(self.laptop_Screen_Size)
            except NoSuchElementException:
                print("Size dropdown not found. Skipping...") 
            except Exception as e:
                print(f"An unexpected error occurred: {str(e)}")        

            # Choose and enter ad price if present
            try:
                WebDriverWait(self.kjj, 60).until(EC.presence_of_element_located((By.ID, "PriceAmount")))
                price = self.kjj.find_element(By.ID,'PriceAmount')
                if self.current_ad_price:
                    price.send_keys(self.current_ad_price)
            except NoSuchElementException:
                print("Price input field not found. Skipping...")
            except Exception as e:
                print(f"An unexpected error occurred: {str(e)}")

            # Enter description if present
            try:
                desc = self.kjj.find_element(By.TAG_NAME,'textarea')
                if self.description:
                    desc.send_keys(self.description)
            except NoSuchElementException:
                print("Description input field not found. Skipping...")
            except Exception as e:
                print(f"An unexpected error occurred: {str(e)}")

            # Upload images if present
            try:
                script_dir = os.path.dirname(os.path.abspath(__file__))

                folder_path = os.path.normpath(os.path.join(script_dir, './Kijiji Ad Photos/', self.current_folderName))
                # folder_path = os.path.abspath(os.path.join('C:/Users/user/Documents/Kijiji Ad Photos', self.current_folderName))
                if folder_path:
                    image_list = os.listdir(folder_path)
                    upload = self.kjj.find_element(By.CLASS_NAME, 'imageUploadButtonWrapper')
                    upload = upload.find_element(By.TAG_NAME, 'input')

                    for pic in image_list:
                        absolute_path = os.path.join(folder_path, pic)  # Construct absolute path to each image
                        print("Absolute path:", absolute_path)
                        upload.send_keys(absolute_path)

                    print("Photos Uploaded")
            except NoSuchElementException:
                print("Image upload button not found. Skipping...")
            except Exception as e:
                print(f"An unexpected error occurred: {str(e)}")

            # Wait for the element with ID "PhoneNumber" to be visible
            try:
                WebDriverWait(self.kjj, 10).until(EC.visibility_of_element_located((By.ID, "PhoneNumber")))
                phone = self.kjj.find_element(By.ID,'PhoneNumber')
                # Once the element is visible, send keys to it
                numbers_list = list(str(self.Phone))
                for number in numbers_list:
                    phone.send_keys(int(number))
                
                print("Phone Number Inserted")
            except NoSuchElementException:
                print("Phone number input field not found. Skipping...")
            except Exception as e:
                print(f"An unexpected error occurred: {str(e)}")
            
            try:
                if self.tags: 
                    tags_list = self.tags.split(',')
                    tags = self.kjj.find_element(By.ID,'pstad-tagsInput')
                
                    for tag in tags_list:
                        tags.send_keys(tag)
                        tags.send_keys(Keys.RETURN)  # Press ENTER key to submit the tag

                    print("Tags Inserted")
            except NoSuchElementException:
                print("Tags input field not found. Skipping...")
            except Exception as e:
                print(f"An unexpected error occurred: {str(e)}")
            
            time.sleep(20)
            try:
                post_ad_link = wait.until(EC.element_to_be_clickable((By.XPATH, '//button[text()="Post Your Ad"]')))
                post_ad_link.click()
            except TimeoutException:
                print("Timeout occurred while waiting for 'Post Your Ad' button. Proceeding without posting the ad.")
            except NoSuchElementException:
                print("Element 'Post Your Ad' button not found. Proceeding without posting the ad.")
            except Exception as e:
                print(f"An unexpected error occurred: {str(e)}")

            print("Post Ad button Clicked!")
            time.sleep(20)        
            
            print(
                'Ad posted successfully.',
                '\nCurrent title is {}.'.format(self.current_ad_title),
                '\nCurrent price is {}.'.format(self.current_ad_price)
            )
        except Exception as e:
            print(f"An unexpected error occurred: {str(e)}")

            self.next_url('https://www.kijiji.ca/m-my-ads/active/1')

    def login(self,credentials):

        self.username = credentials['username']
        self.password = credentials['password']
        time.sleep(5)
        while True:
            try:
                WebDriverWait(self.kjj, 60).until(EC.presence_of_element_located((By.ID, "username")))
                user = self.kjj.find_element(By.ID,'username')
                user.send_keys(self.username)
                
                WebDriverWait(self.kjj, 60).until(EC.presence_of_element_located((By.ID, "password")))
                password = self.kjj.find_element(By.ID,'password')
                password.send_keys(self.password)
                                
                WebDriverWait(self.kjj, 60).until(EC.presence_of_element_located((By.ID, "login-submit")))
                self.next_click(self.kjj.find_element(By.ID, "login-submit"))
                break
            except TimeoutException:
                print("TimeoutException: One or more elements did not load within 60 seconds.")
            # Navigate back using browser navigation commands
                self.kjj.back()
                print("Navigated back to the previous page.")
            except Exception as e:
                print(f"An unexpected error occurred: {str(e)}")
    
    def delete_ad(self,ad_data):
        wait = WebDriverWait(self.kjj, 60)
        self.current_ad_title = ad_data['Title'].strip()
        print(self.current_ad_title)
        self.next_url('https://www.kijiji.ca/m-my-ads/active/1?siteLocale=en_CA')
        time.sleep(5)   
        div_elements = []
        while True:
            try:
                tbody_element = self.kjj.find_element(By.TAG_NAME, 'tbody')
                tr_elements = tbody_element.find_elements(By.TAG_NAME,'tr')

                for tr_element in tr_elements:
                    div_elements = tr_element.find_elements(By.XPATH,'.//*')   
                    found_ad_to_delete = False
                    if not div_elements:
                        print("No div elements found.")
                    else:        
                        for div_element in div_elements:
                            links = div_element.find_elements(By.XPATH, ".//a[@href]")
                            for link in links:
                                if self.current_ad_title in link.get_attribute("innerHTML"):    
                                    try:
                                        delete_button = tr_element.find_element(By.XPATH, './/td[8]/div[1]/div[2]/div/div[2]/button/span')
                                        delete_button.click()
                                        found_ad_to_delete = True
                                        break
                                    except NoSuchElementException:
                                        print("Delete button not found for the ad.")
                                    except Exception as e:
                                        print(f"An unexpected error occurred: {str(e)}")
                            if found_ad_to_delete:
                                try:
                                    button = WebDriverWait(self.kjj, 10).until(EC.element_to_be_clickable((By.XPATH, '//button[text()="Prefer not to say"]')))
                                    button.click()
                                    time.sleep(2)
                                    button = WebDriverWait(self.kjj, 10).until(EC.element_to_be_clickable((By.XPATH, '//button[text()="Delete My Ad"]')))
                                    button.click()
                                    time.sleep(2)
                                    print('Ad deleted.')
                                    button = WebDriverWait(self.kjj, 10).until(EC.element_to_be_clickable((By.ID, 'modalCloseButton')))
                                    button.click()
                                    time.sleep(2)
                                    self.kjj.refresh()
                                    time.sleep(10)
                                    return
                                except TimeoutException:
                                    print('Timeout occurred while deleting the ad.')
                                except Exception as e:
                                    print(f"An unexpected error occurred: {str(e)}")
                            else:
                                print('Ad with title "{}" not found for deletion.'.format(self.current_ad_title))

                next_active_page = self.kjj.find_element(By.XPATH, '/html/body/div[3]/div[4]/div/div/div/div/div[3]/a')
                if next_active_page.text.strip() == "Next":
                    next_active_page.click()
                else:
                    return
            
            except StaleElementReferenceException:
                # Handle stale element exception by re-finding the element
                print("Stale element reference. Retrying...")
                continue

            except NoSuchElementException:
                print("Element not found.")
                return
                # Handle the situation here, such as returning from the method or raising an error

            except Exception as e:
                print(f"An unexpected error occurred: {str(e)}")
                return
            

    def check_Ads(self,ad_data):
        wait = WebDriverWait(self.kjj, 60)
        self.current_ad_title = ad_data['Title']
        self.next_url('https://www.kijiji.ca/m-my-ads/active/1')
        time.sleep(5)   
        div_elements = []        
        while True:
            try:
                tbody_element = self.kjj.find_element(By.TAG_NAME, 'tbody')
                tr_elements = tbody_element.find_elements(By.TAG_NAME,'tr')

                for tr_element in tr_elements:
                    div_elements = tr_element.find_elements(By.XPATH,'.//*')   
                    found_ad_to_delete = False
                    if not div_elements:
                        print("No div elements found.")
                    else:        
                        for div_element in div_elements:
                            links = div_element.find_elements(By.XPATH, ".//a[@href]")
                            for link in links:
                                if self.current_ad_title in link.get_attribute("innerHTML"): 
                                    return 0

                next_active_page = self.kjj.find_element(By.XPATH, '/html/body/div[3]/div[4]/div/div/div/div/div[3]/a')
                if next_active_page.text.strip() == "Next":
                    next_active_page.click()
                else:
                    return
            
            except StaleElementReferenceException:
                # Handle stale element exception by re-finding the element
                print("Stale element reference. Retrying...")
                continue

            except NoSuchElementException:
                print("Element not found.")
                # Handle the situation here, such as returning from the method or raising an error
            except Exception as e:
                print(f"An unexpected error occurred: {str(e)}")


    def read_txt(self,file_path):
        credentials = {}
        with open(file_path, 'r') as file:
            for line in file:
                key, value = line.strip().split(':')
                credentials[key] = value
        return credentials


        
            

def main():
    for j in range (1,2):


        # for i in range(1, 480):  # Start from 1 and end at 720
        #     print(f"Iteration {i}")
        #     time.sleep(60)
        browser = kijiji()
        file_path = 'credentials.txt'
        credentials = browser.read_txt(file_path)
        json_data = browser.connect_db(credentials)
        # ad_data_list = browser.fetch_data()
        # print(ad_data_list) 

        
        

        browser.access_kijiji(credentials)

        try:
        # Example usage
            # source_file = "Master Excel.xlsx"
            # source_sheet_index = credentials['sheetNumber']  # Index of the second sheet (1-based)
            # destination_file = "ads_data_new.xlsx"

            # browser.copy_sheet(source_file, source_sheet_index, destination_file)
            # Load data from Excel file
            print("About to Start")
            # wb = load_workbook('ads_data_new.xlsx')
            print("Ads_Data Loaded")
            # sheet = wb.active
            # print(json_data)
            # Start from the second row (assuming the first row is headers)
            # for row in sheet.iter_rows(min_row=2, values_only=True):
            data = json.loads(json_data)
            print(data)
            # Iterate over rows
            for row in data:
                print("Processing row:", row)    
                ad_data = {
                    'Title': row.get('Title'),
                    'Category': row.get('Category'),
                    'Price': row.get('Price'),
                    'Description': row.get('Description'),
                    'Condition': row.get('Condition'),
                    'PhoneBrand': row.get('PhoneBrand'),
                    'PhoneBrandCarrier': row.get('PhoneBrandCarrier'),
                    'Images_FolderName': row.get('Images_FolderName'),
                    'Phone': row.get('Phone'),
                    'Tags': row.get('Tags'),
                    'Size': row.get('Size'),
                    'Type': row.get('Type'),
                    'Tablet Brand': row.get('Tablet Brand'),
                    'laptop Screen Size': row.get('laptop Screen Size')
                }
            
                print(ad_data)
                browser.delete_ad(ad_data)
                print("Ad Deleted")
                time.sleep(10)
                
                browser.post_ad(ad_data)
                print("Ad Posted")
                time.sleep(30)  # Sleep between posting ads

            # for row in sheet.iter_rows(min_row=2, values_only=True):
            #     ad_data = {
            #         'Title': row[0],
            #         'Category': row[1],
            #         'Price': row[2],
            #         'Description': row[3],
            #         'Condition': row[4],
            #         'PhoneBrand': row[5],
            #         'PhoneBrandCarrier': row[6],
            #         'Images_FolderName': row[7],
            #         'Phone': row[8],
            #         'Tags': row[9], 
            #         'Size': row[10],
            #         'Type': row[11],
            #         'Tablet Brand': row[12],
            #         'laptop Screen Size': row[13]
            #     }
            #     if browser.check_Ads(ad_data) != 0:
            #         browser.post_ad(ad_data) 
            #         print("Deleted Ad Posted")

                time.sleep(30)  # Sleep between posting ads
                
                # account_button = WebDriverWait(browser, 10).until(
                #     EC.element_to_be_clickable((By.CSS_SELECTOR, 'button[aria-label="My Account"]'))
                # )
                # account_button.click()

                # # 2. Wait for the dropdown menu to appear and become interactive
                # dropdown_menu = WebDriverWait(browser, 10).until(
                #     EC.visibility_of_element_located((By.CLASS_NAME, 'root-3161363123'))
                # )

                # # 3. Locate and click the "Log Out" button within the dropdown
                # logout_button = WebDriverWait(dropdown_menu, 10).until(
                #     EC.element_to_be_clickable((By.CSS_SELECTOR, 'button[data-testid="header-logout"]'))
                # )
                # logout_button.click()
        
        except Exception as e:
            print(f"Error occurred: {e}")

        # for i in range(1, 480):  # Start from 1 and end at 720
        #     print(f"Iteration {i}")
        #     time.sleep(60)

if __name__ == '__main__':
    main()
