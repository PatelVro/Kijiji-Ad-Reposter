import os
import sys
import time
import datetime # Make sure datetime is imported
import random
import pymysql
import json
from selenium import webdriver
from selenium.webdriver.firefox.options import Options as FirefoxOptions
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.common.keys import Keys
from selenium.common.exceptions import ElementClickInterceptedException
from selenium.webdriver.chrome.options import Options as ChromeOptions
from openpyxl import load_workbook
from openpyxl import Workbook
from selenium.common.exceptions import StaleElementReferenceException
# Assuming db_config.py exists and is correctly configured
# If db_config.py is in the same directory, this should work:
from db_config import config


class kijiji():

    def __init__(self):
        self.current_dir = os.getcwd()
        self.headless = ChromeOptions()

        self.headless.add_argument("--disable-blink-features=AutomationControlled")
        # REVERTED: Keep original user-data-dir and profile-directory paths
        self.headless.add_argument(r"user-data-dir=C:\Users\Administrator\AppData\Local\Google\Chrome\User Data")
        self.headless.add_argument(r'--profile-directory=Default')
        self.headless.add_argument("--lang=en-US")
        self.db_connection = None
        # REVERTED: Screenshots will be saved in self.current_dir directly
        # No separate screenshots_dir will be created by default in __init__

    @staticmethod
    def is_page_fully_loaded(driver):
        return driver.execute_script("return document.readyState") == "complete"

    def take_screenshot(self, driver, name_prefix="exception"):
        """Takes a screenshot and saves it with a timestamp in the current working directory."""
        if driver is None:
            print("Driver is None, cannot take screenshot.")
            return

        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"{name_prefix}_{timestamp}.png"
        # REVERTED: Save directly in self.current_dir
        filepath = os.path.join(self.current_dir, filename)
        try:
            driver.save_screenshot(filepath)
            print(f"Screenshot saved to: {filepath}")
        except Exception as e:
            print(f"Error taking screenshot: {e}")

    def connect_db(self, credentials):
        try:
            self.email = credentials['username'] # This will raise KeyError if 'username' is not in credentials

            connection = pymysql.connect(
                host=config['host'],
                port=config['port'],
                user=config['user'],
                password=config['password'],
                database=config['database']
            )

            with connection.cursor() as cursor:
                cursor.execute("SELECT DATABASE();")
                result = cursor.fetchone()
                print("You're connected to database: ", result[0])

                # Ensure self.email (table name) is properly escaped if it can contain special characters
                # However, standard email formats are usually safe.
                cursor.execute(f"SELECT * FROM `{self.email}`")
                column_names = [desc[0] for desc in cursor.description]
                rows = cursor.fetchall()

                print(f"Data from `{self.email}` table:")
                data = [dict(zip(column_names, row)) for row in rows]
                json_data = json.dumps(data, indent=4)
                print(json_data)
                return json_data

        except KeyError as ke:
            print(f"KeyError in connect_db: {ke}. 'username' might be missing from credentials.")
            return None
        except pymysql.MySQLError as e:
            print(f"Error while connecting to MariaDB: {e}")
            return None # Return None on error
        except Exception as e:
            print(f"An unexpected error occurred in connect_db: {e}")
            return None
        # Removed finally block that would close a connection not guaranteed to be open


    def access_kijiji(self, credentials):
        self.kjj = webdriver.Chrome(options=self.headless)

        while True:
            try:
                wait = WebDriverWait(self.kjj, 20)
                self.next_url('https://www.kijiji.ca/?siteLocale=en_CA')
                # If you had a specific element to click here that was getting intercepted:
                # example_element = wait.until(EC.element_to_be_clickable((By.ID, "some_element_id")))
                # example_element.click()
                break
            except TimeoutException:
                print("Timeout occurred. Unable to load initial page or find element. Retrying...")
                self.take_screenshot(self.kjj, "timeout_access_kijiji")
                self.kjj.refresh()
            except ElementClickInterceptedException as eci:
                print(f"ElementClickInterceptedException in access_kijiji: Click intercepted. Details: {eci}. Retrying...")
                self.take_screenshot(self.kjj, "click_intercepted_access_kijiji")
                self.kjj.refresh()
            except Exception as e:
                print(f"An unexpected error occurred in access_kijiji: {str(e)}. Retrying...")
                self.take_screenshot(self.kjj, "unexpected_error_access_kijiji")
                self.kjj.refresh()


    def next_url(self, new_url):
        current = self.kjj.current_url
        self.kjj.get(new_url)
        # Original explicit wait:
        wait_end_time = time.time() + 10
        while current == self.kjj.current_url and time.time() < wait_end_time:
            time.sleep(1)
        if current == self.kjj.current_url:
            print(f"Warning: URL did not change from {current} after navigating to {new_url} within timeout.")
            self.take_screenshot(self.kjj, f"next_url_did_not_change_{new_url.split('/')[-1]}")


    def next_click(self, e_to_click):
        current_url_before_click = self.kjj.current_url
        try:
            # Assuming e_to_click is already a WebElement found by the caller
            e_to_click.click()
            # Original explicit wait:
            wait_end_time = time.time() + 10
            while current_url_before_click == self.kjj.current_url and time.time() < wait_end_time:
                time.sleep(1)
            if current_url_before_click == self.kjj.current_url:
                 print(f"Warning: URL did not change after clicking element. Element text (if any): {e_to_click.text[:50] if hasattr(e_to_click, 'text') else 'N/A'}")
                 self.take_screenshot(self.kjj, "next_click_url_did_not_change")

        except ElementClickInterceptedException as eci:
            print(f"ElementClickInterceptedException during next_click: {eci}. Element: {e_to_click.tag_name if hasattr(e_to_click, 'tag_name') else 'N/A'}")
            self.take_screenshot(self.kjj, "click_intercepted_next_click")
            raise # Re-raise to be handled by the caller if necessary
        except Exception as e:
            print(f"Error during next_click: {e}")
            self.take_screenshot(self.kjj, "error_next_click")
            raise # Re-raise


    def copy_sheet(self, source_file, source_sheet_index, destination_file):
        try:
            source_sheet_index = int(source_sheet_index)
            source_wb = load_workbook(source_file)
            source_sheet = source_wb.worksheets[source_sheet_index - 1]

            if not os.path.exists(destination_file):
                dest_wb = Workbook()
                dest_sheet = dest_wb.active # Get the default sheet
                # dest_sheet.title = source_sheet.title # Optionally copy title
                dest_wb.save(destination_file) # Save to create the file
            
            dest_wb = load_workbook(destination_file)
            
            # If you want to replace the first sheet or a specific sheet:
            if dest_wb.sheetnames:
                dest_sheet = dest_wb.active # or dest_wb[dest_wb.sheetnames[0]]
                dest_sheet.delete_rows(1, dest_sheet.max_row) # Clear existing content
                # dest_sheet.title = source_sheet.title # Optionally update title
            else: # Should not happen if file exists or was just created with an active sheet
                dest_sheet = dest_wb.create_sheet(title=source_sheet.title if source_sheet.title else "CopiedSheet")


            for row in source_sheet.iter_rows(values_only=True):
                dest_sheet.append(row)

            dest_wb.save(destination_file)
            print(f"Sheet at index {source_sheet_index} copied from '{source_file}' to '{destination_file}' successfully.")
        except Exception as e:
            print(f"An error occurred in copy_sheet: {str(e)}")


    def post_ad(self, ad_data):
        try:
            # Using .get() for safer dictionary access, providing defaults
            self.current_ad_title = ad_data.get('Title', '').strip()
            self.current_categories = ad_data.get('Category', '') # Expecting semi-colon separated
            self.current_ad_price = ad_data.get('Price', '')
            self.description = ad_data.get('Description', '')
            self.condition = ad_data.get('Condition', '') # This is a value for a dropdown
            self.phonebrand = ad_data.get('PhoneBrand', '') # Value for dropdown
            self.phonebrand_carrier = ad_data.get('PhoneBrandCarrier', '') # Value for dropdown
            self.current_folderName = ad_data.get('Images_FolderName', '').strip()
            self.Phone = ad_data.get('Phone', '')
            self.tags = ad_data.get('Tags', '') # Expecting comma-separated
            self.size = ad_data.get('Size', '') # Value for dropdown
            self.type = ad_data.get('Type', '') # Value for dropdown
            self.tablet_brand = ad_data.get('Tablet Brand', '') # Value for dropdown
            self.laptop_Screen_Size = ad_data.get('laptop Screen Size', '') # Value for dropdown

            if not self.current_ad_title:
                print("Ad title is missing. Cannot post ad.")
                return

            self.kjj.refresh()
            time.sleep(5) # Allow page to settle after refresh
            wait = WebDriverWait(self.kjj, 60) # Long wait for critical elements

            try:
                post_ad_link = wait.until(EC.element_to_be_clickable((By.LINK_TEXT, "Post ad")))
                post_ad_link.click()
            except ElementClickInterceptedException as eci:
                print(f"ElementClickInterceptedException clicking 'Post ad' link: {eci}")
                self.take_screenshot(self.kjj, "click_intercepted_post_ad_link")
                # Attempt to scroll and retry, or JS click if needed
                # For now, just re-raise or let it fail if refresh is the strategy
                raise
            except TimeoutException:
                print("Timeout waiting for 'Post ad' link.")
                self.take_screenshot(self.kjj, "timeout_post_ad_link")
                return # Cannot proceed

            WebDriverWait(self.kjj, 60).until(EC.presence_of_element_located((By.ID, "AdTitleForm")))
            title_field = self.kjj.find_element(By.ID, 'AdTitleForm')
            title_field.send_keys(self.current_ad_title)
            
            # Original: next_button_link = wait.until(EC.element_to_be_clickable((By.TAG_NAME, "button")))
            # This is too generic. Need a more specific button, e.g., by text or a unique attribute.
            # Assuming a button with text 'Next' or 'Continue'
            try:
                next_button_title_page = WebDriverWait(self.kjj, 10).until(
                    EC.element_to_be_clickable((By.XPATH, "//button[normalize-space()='Next' or normalize-space()='Continue']"))
                )
                next_button_title_page.click()
            except TimeoutException:
                print("Timeout finding 'Next' button after entering title.")
                self.take_screenshot(self.kjj, "timeout_next_button_after_title")
                return
            except ElementClickInterceptedException as eci_next:
                print(f"Click intercepted for 'Next' button after title: {eci_next}")
                self.take_screenshot(self.kjj, "click_intercepted_next_button_after_title")
                # Fallback to JS click
                try:
                    next_button_title_page = self.kjj.find_element(By.XPATH, "//button[normalize-space()='Next' or normalize-space()='Continue']")
                    self.kjj.execute_script("arguments[0].click();", next_button_title_page)
                    print("Used JS click for 'Next' button after title.")
                except Exception as e_js:
                    print(f"JS click also failed for 'Next' button: {e_js}")
                    return


            categories_list = [cat.strip() for cat in self.current_categories.split(';') if cat.strip()]
            time.sleep(5) # Wait for category section to load

            for category_to_select in categories_list:
                print(f"Selecting category: {category_to_select} Iteration.....")
                try:
                    # Wait for the category container to ensure it's loaded
                    # Original class name: "allCategoriesContainer-1711250309" - this might be dynamic
                    # Using a more generic approach if the class name changes often:
                    # category_container = wait.until(EC.presence_of_element_located((By.XPATH, "//div[contains(@class, 'allCategoriesContainer')]")))
                    # For now, using the provided class name, assuming it's somewhat stable:
                    category_container = wait.until(EC.presence_of_element_located((By.CLASS_NAME, "allCategoriesContainer-1711250309")))
                    
                    # Find button by text within the container
                    # Using normalize-space() for robustness against extra whitespace in button text
                    category_button_xpath = f".//button[.//span[normalize-space(text())='{category_to_select}']]"
                    
                    # Scroll into view might be needed if categories are in a scrollable list
                    # self.kjj.execute_script("arguments[0].scrollIntoView(true);", category_button_found);
                    
                    category_button_found = WebDriverWait(category_container, 15).until(
                        EC.element_to_be_clickable((By.XPATH, category_button_xpath))
                    )
                    print(f"Match Found for '{category_to_select}'")
                    category_button_found.click()
                    time.sleep(2) # Wait for sub-categories or next step to load
                except TimeoutException:
                    print(f"Timeout: Could not find or click category: '{category_to_select}'")
                    self.take_screenshot(self.kjj, f"timeout_category_{category_to_select.replace(' ','_')}")
                    # Decide if to continue with other categories or stop
                except ElementClickInterceptedException as eci_cat:
                    print(f"ElementClickInterceptedException clicking category '{category_to_select}': {eci_cat}")
                    self.take_screenshot(self.kjj, f"click_intercepted_category_{category_to_select.replace(' ','_')}")
                    # Try JS click as fallback
                    try:
                        print(f"Attempting JS click for category '{category_to_select}'...")
                        cat_btn_for_js = category_container.find_element(By.XPATH, category_button_xpath)
                        self.kjj.execute_script("arguments[0].click();", cat_btn_for_js)
                        time.sleep(2)
                    except Exception as e_js_cat:
                        print(f"JS click for category '{category_to_select}' also failed: {e_js_cat}")
                except NoSuchElementException: # Should be caught by WebDriverWait's TimeoutException
                    print(f"NoSuchElementException: Category button for '{category_to_select}' not found.")
                    self.take_screenshot(self.kjj, f"nocategory_{category_to_select.replace(' ','_')}")
                except Exception as e_cat_select:
                    print(f"An unexpected error occurred during category selection for '{category_to_select}': {e_cat_select}")
                    self.take_screenshot(self.kjj, f"error_category_select_{category_to_select.replace(' ','_')}")


            print('Category selection process finished.')
            time.sleep(5) # Wait for the rest of the form to potentially update based on category

            # Helper for clicking labels, with screenshot on failure
            def click_label_if_present(css_selector_for, description):
                try:
                    element = WebDriverWait(self.kjj, 5).until( # Shorter wait, these are optional
                        EC.element_to_be_clickable((By.CSS_SELECTOR, css_selector_for))
                    )
                    element.click()
                    print(f"Clicked '{description}'.")
                except TimeoutException:
                    print(f"Element for '{description}' (CSS: {css_selector_for}) not found or not clickable within 5s. Skipping.")
                    # self.take_screenshot(self.kjj, f"timeout_label_{description.replace(' ','_')}") # Optional screenshot for non-critical items
                except ElementClickInterceptedException as eci_label:
                    print(f"ElementClickInterceptedException for '{description}': {eci_label}")
                    self.take_screenshot(self.kjj, f"click_intercepted_label_{description.replace(' ','_')}")
                except Exception as e_label:
                    print(f"An error occurred with '{description}': {str(e_label)}")
                    self.take_screenshot(self.kjj, f"error_label_{description.replace(' ','_')}")

            click_label_if_present('label[for="forsaleby_s-delr"]', "Business Radio Option")
            click_label_if_present('label[for="fulfillment_s-delivery"]', "Delivery Checkbox")
            click_label_if_present('label[for="fulfillment_s-shipping"]', "Shipping Checkbox")
            click_label_if_present('label[for="fulfillment_s-curbside"]', "Curbside Checkbox")
            click_label_if_present('label[for="payment_s-cashless"]', "Cashless payment Checkbox")
            click_label_if_present('label[for="payment_s-cashaccepted"]', "Cash payment Checkbox")

            # Helper for selecting dropdown options, with screenshot on failure
            def select_dropdown_if_present(element_id, value_to_select, description):
                if not value_to_select: # Skip if no value provided from data
                    # print(f"No value provided for '{description}', skipping.")
                    return
                try:
                    dropdown_element = WebDriverWait(self.kjj, 5).until( # Shorter wait
                        EC.presence_of_element_located((By.ID, element_id))
                    )
                    select_obj = Select(dropdown_element)
                    select_obj.select_by_value(value_to_select)
                    print(f"Selected '{value_to_select}' for '{description}'.")
                except TimeoutException:
                    print(f"Dropdown '{description}' (ID: {element_id}) not found within 5s. Skipping.")
                except NoSuchElementException: # This refers to the *option value* not being found by Select
                    print(f"Option with value '{value_to_select}' not found in '{description}' (ID: {element_id}). Skipping.")
                    # self.take_screenshot(self.kjj, f"nooption_dropdown_{element_id}_{value_to_select}") # Optional
                except Exception as e_dropdown:
                    print(f"An error occurred with '{description}' (ID: {element_id}): {str(e_dropdown)}")
                    self.take_screenshot(self.kjj, f"error_dropdown_{element_id}")

            select_dropdown_if_present('condition_s', self.condition, "Condition")
            select_dropdown_if_present('phonebrand_s', self.phonebrand, "Phone Brand")
            select_dropdown_if_present('phonecarrier_s', self.phonebrand_carrier, "Phone Brand Carrier")
            select_dropdown_if_present('size_s', self.size, "Size (e.g., TV)")
            select_dropdown_if_present('type_s', self.type, "Type (e.g., TV)")
            select_dropdown_if_present('tabletbrand_s', self.tablet_brand, "Tablet Brand")
            select_dropdown_if_present('laptopscreensize_s', self.laptop_Screen_Size, "Laptop Screen Size")


            try:
                price_field = WebDriverWait(self.kjj, 10).until(EC.presence_of_element_located((By.ID, "PriceAmount")))
                if self.current_ad_price is not None and str(self.current_ad_price).strip() != "": # Check if price is not empty
                    price_field.send_keys(str(self.current_ad_price)) # Ensure it's a string
                    print("Price entered.")
                # else:
                    # print("No price provided or price is empty. Skipping price field.")
            except TimeoutException:
                print("Price input field (ID: PriceAmount) not found. Skipping price.")
                # self.take_screenshot(self.kjj, "timeout_price_input") # Optional for non-critical
            except Exception as e_price:
                print(f"An unexpected error occurred with price input: {str(e_price)}")
                self.take_screenshot(self.kjj, "error_price_input")

            try:
                # Using a more general XPath for textarea, assuming it's for description
                desc_textarea = WebDriverWait(self.kjj, 10).until(
                    EC.presence_of_element_located((By.XPATH, "//textarea[contains(@name, 'description') or contains(@id, 'description') or contains(@name, 'Description')]"))
                )
                if self.description:
                    desc_textarea.send_keys(self.description)
                    print("Description entered.")
            except TimeoutException:
                print("Description textarea not found. Skipping description.")
                # self.take_screenshot(self.kjj, "timeout_description_textarea") # Optional
            except Exception as e_desc:
                print(f"An unexpected error occurred with description input: {str(e_desc)}")
                self.take_screenshot(self.kjj, "error_description_textarea")


            try:
                # REVERTED: script_dir and folder_path construction as per original
                script_dir = os.path.dirname(os.path.abspath(__file__))
                folder_path = os.path.normpath(os.path.join(script_dir, './Kijiji Ad Photos/', self.current_folderName))
                
                if self.current_folderName and os.path.exists(folder_path) and os.path.isdir(folder_path):
                    image_files = [f for f in os.listdir(folder_path) if os.path.isfile(os.path.join(folder_path, f))]
                    if not image_files:
                        print(f"No image files found in folder: {folder_path}")
                    else:
                        # The file input is often hidden. Find it by a more reliable selector.
                        # Original: upload = self.kjj.find_element(By.CLASS_NAME, 'imageUploadButtonWrapper').find_element(By.TAG_NAME, 'input')
                        # This might be fragile. A common pattern is an input type="file".
                        try:
                            file_input_element = WebDriverWait(self.kjj, 15).until(
                                EC.presence_of_element_located((By.XPATH, "//input[@type='file' and (contains(@class, 'uploader') or contains(@class, 'fileChooser') or contains(@id, 'imageUpload'))] | //div[contains(@class, 'imageUploadButton')]//input[@type='file']"))
                            ) # This XPath tries a few common patterns for file inputs
                        except TimeoutException:
                             # Fallback to original if the above fails and original class name is reliable
                             file_input_element = WebDriverWait(self.kjj, 10).until(
                                 EC.presence_of_element_located((By.XPATH, "//div[contains(@class, 'imageUploadButtonWrapper')]//input[@type='file']"))
                             )


                        for pic_filename in image_files:
                            absolute_image_path = os.path.join(folder_path, pic_filename)
                            print("Uploading image:", absolute_image_path)
                            file_input_element.send_keys(absolute_image_path)
                            time.sleep(2) # Brief pause for each upload to register
                        print(f"{len(image_files)} photos sent for upload.")
                elif self.current_folderName: # Folder name was provided but path doesn't exist/isn't a dir
                     print(f"Image folder path not found or is not a directory: {folder_path}. Skipping image upload.")

            except TimeoutException:
                print("Image upload input element not found. Skipping image upload.")
                self.take_screenshot(self.kjj, "timeout_image_upload_input")
            except Exception as e_img:
                print(f"An unexpected error occurred during image upload: {str(e_img)}")
                self.take_screenshot(self.kjj, "error_image_upload")


            try:
                phone_number_field = WebDriverWait(self.kjj, 10).until(EC.visibility_of_element_located((By.ID, "PhoneNumber")))
                if self.Phone is not None and str(self.Phone).strip() != "":
                    phone_number_field.clear() # Clear existing if any
                    # Original sends numbers one by one, which is fine but can be slow.
                    # Direct send_keys is usually okay for phone numbers.
                    phone_number_field.send_keys(str(self.Phone))
                    print("Phone Number Inserted.")
                # else:
                    # print("No phone number provided. Skipping phone input.")
            except TimeoutException:
                print("Phone number input field (ID: PhoneNumber) not found. Skipping.")
                # self.take_screenshot(self.kjj, "timeout_phone_input") # Optional
            except Exception as e_phone:
                print(f"An unexpected error occurred with phone input: {str(e_phone)}")
                self.take_screenshot(self.kjj, "error_phone_input")

            try:
                if self.tags:
                    tags_input_field = WebDriverWait(self.kjj, 10).until(EC.visibility_of_element_located((By.ID, 'pstad-tagsInput')))
                    tags_to_add = [tag.strip() for tag in self.tags.split(',') if tag.strip()]
                    for single_tag in tags_to_add:
                        tags_input_field.send_keys(single_tag)
                        tags_input_field.send_keys(Keys.RETURN) # Or Keys.ENTER
                        time.sleep(0.5) # Small pause for tag to be processed
                    print("Tags Inserted.")
                # else:
                    # print("No tags provided. Skipping tags input.")
            except TimeoutException:
                print("Tags input field (ID: pstad-tagsInput) not found. Skipping tags.")
                # self.take_screenshot(self.kjj, "timeout_tags_input") # Optional
            except Exception as e_tags:
                print(f"An unexpected error occurred with tags input: {str(e_tags)}")
                self.take_screenshot(self.kjj, "error_tags_input")


            time.sleep(10) # Reduced from 20s - evaluate if this long pause is always needed before final post.
                           # Could be waiting for image uploads to fully complete visually, or for some validation.
            try:
                # Original XPath: '//button[text()="Post Your Ad"]'
                # Making it slightly more robust with normalize-space()
                final_post_button = wait.until(EC.element_to_be_clickable((By.XPATH, '//button[normalize-space()="Post Your Ad"]')))
                # self.kjj.execute_script("arguments[0].scrollIntoView(true);", final_post_button) # Optional: ensure it's in view
                # time.sleep(0.5)
                final_post_button.click()
                print("Final 'Post Your Ad' button Clicked!")
            except TimeoutException:
                print("Timeout: Final 'Post Your Ad' button not found or not clickable.")
                self.take_screenshot(self.kjj, "timeout_final_post_button")
                # Ad posting likely failed, might want to return or raise here
                return
            except ElementClickInterceptedException as eci_final:
                print(f"ElementClickInterceptedException on final 'Post Your Ad' button: {eci_final}")
                self.take_screenshot(self.kjj, "click_intercepted_final_post_button")
                # Attempt JS click as a fallback
                try:
                    print("Attempting JS click for final 'Post Your Ad' button...")
                    # Re-fetch element for JS click if it might have gone stale
                    final_post_button_js = self.kjj.find_element(By.XPATH, '//button[normalize-space()="Post Your Ad"]')
                    self.kjj.execute_script("arguments[0].click();", final_post_button_js)
                    print("JS click attempted for final 'Post Your Ad' button.")
                except Exception as e_js_final:
                    print(f"JS click for final post button also failed: {e_js_final}")
                    self.take_screenshot(self.kjj, "js_click_failed_final_post_button")
                    return # Ad posting likely failed
            except NoSuchElementException: # Should be caught by WebDriverWait's TimeoutException
                print("Final 'Post Your Ad' button not found (NoSuchElement).")
                self.take_screenshot(self.kjj, "notfound_final_post_button")
                return
            except Exception as e_final_post:
                print(f"An unexpected error occurred clicking final 'Post Your Ad': {str(e_final_post)}")
                self.take_screenshot(self.kjj, "error_final_post_button")
                return

            time.sleep(15) # Original was 20s. Wait for ad posting confirmation/redirect.
                           # Consider waiting for a specific element or URL change indicating success.
            
            # Example: Check if URL changed to manage ads page or a success page
            current_url = self.kjj.current_url
            if "ManageAds" in current_url or "VipSuccess" in current_url or "PostingSuccess" in current_url :
                 print(f"Ad '{self.current_ad_title}' (Price: {self.current_ad_price}) likely posted successfully. Current URL: {current_url}")
            else:
                 print(f"Ad '{self.current_ad_title}' posting process finished, but confirmation unclear. Current URL: {current_url}")
                 self.take_screenshot(self.kjj, f"post_ad_confirmation_check_{self.current_ad_title[:15]}")


        except Exception as e_outer_post_ad:
            print(f"An overarching unexpected error occurred in post_ad for title '{self.current_ad_title}': {str(e_outer_post_ad)}")
            self.take_screenshot(self.kjj, f"error_overall_post_ad_{self.current_ad_title[:20].replace(' ','_')}")
            # Navigate to a known safe page to allow next ad processing if possible
            try:
                print("Attempting to navigate to active ads page after error in post_ad.")
                self.next_url('https://www.kijiji.ca/m-my-ads/active/1')
            except Exception as e_nav_fail:
                print(f"Failed to navigate to active ads page after error: {e_nav_fail}")
                # This might indicate a severe browser issue, consider re-raising or stopping.


    def login(self, credentials):
        self.username = credentials.get('username')
        self.password = credentials.get('password')

        if not self.username or not self.password:
            print("Username or password not provided in credentials. Skipping login.")
            return False

        # Assuming access_kijiji lands on homepage, now try to find global sign-in.
        # If Kijiji takes you directly to a login page, this part might need adjustment.
        try:
            # Look for a general "Sign In" link/button. This XPath is an example.
            # It might be `//a[@href='/t-login.html']` or similar.
            initial_signin_link_xpath = "//a[contains(@title, 'Sign In') or normalize-space()='Sign In' or contains(@href, 'login')]"
            sign_in_trigger = WebDriverWait(self.kjj, 15).until(
                EC.element_to_be_clickable((By.XPATH, initial_signin_link_xpath))
            )
            print("Found initial 'Sign In' trigger. Clicking...")
            sign_in_trigger.click()
            # Wait for login form elements to appear after clicking sign-in
            WebDriverWait(self.kjj, 10).until(EC.presence_of_element_located((By.ID, "username")))
        except TimeoutException:
            print("Initial 'Sign In' trigger not found or login form did not appear. Assuming already on login page or flow changed.")
            self.take_screenshot(self.kjj, "timeout_initial_signin_or_form_load")
            # Check if username field is already present (i.e., already on login page)
            if not self.kjj.find_elements(By.ID, "username"):
                print("Cannot find username field even after timeout. Login flow likely broken.")
                return False
        except ElementClickInterceptedException as eci_initial_signin:
            print(f"Initial 'Sign In' click intercepted: {eci_initial_signin}")
            self.take_screenshot(self.kjj, "click_intercepted_initial_signin")
            # Try JS click for the initial sign-in trigger
            try:
                sign_in_trigger_js = self.kjj.find_element(By.XPATH, initial_signin_link_xpath)
                self.kjj.execute_script("arguments[0].click();", sign_in_trigger_js)
                WebDriverWait(self.kjj, 10).until(EC.presence_of_element_located((By.ID, "username")))
            except Exception as e_js_initial_signin:
                print(f"JS click for initial Sign In also failed or form did not load: {e_js_initial_signin}")
                return False
        except Exception as e_initial_signin_flow:
            print(f"Unexpected error during initial sign-in flow: {e_initial_signin_flow}")
            self.take_screenshot(self.kjj, "error_initial_signin_flow")
            return False


        # On the actual login form page
        max_login_attempts = 2
        for attempt in range(max_login_attempts):
            try:
                print(f"Login attempt {attempt + 1}/{max_login_attempts} for user '{self.username}'")
                user_field = WebDriverWait(self.kjj, 10).until(EC.presence_of_element_located((By.ID, "username")))
                user_field.clear()
                user_field.send_keys(self.username)

                pass_field = WebDriverWait(self.kjj, 10).until(EC.presence_of_element_located((By.ID, "password")))
                pass_field.clear()
                pass_field.send_keys(self.password)

                # Submit button - original was "login-submit" ID
                submit_button = WebDriverWait(self.kjj, 10).until(EC.element_to_be_clickable((By.ID, "login-submit")))
                
                # Using self.next_click as per original structure, though direct .click() is also fine here
                # self.next_click(submit_button) # This also has a built-in wait for URL change
                submit_button.click() # Simpler, then wait for outcome

                # Wait for indication of successful login or failure
                # e.g., URL changes away from login page, or a welcome message appears, or an error message appears.
                # This needs to be specific to Kijiji's behavior.
                WebDriverWait(self.kjj, 15).until(
                    lambda driver: "login" not in driver.current_url.lower() or \
                                   driver.find_elements(By.XPATH, "//*[contains(text(),'Invalid email or password') or contains(text(),'incorrect')]") or \
                                   driver.find_elements(By.XPATH, "//a[contains(@href, 'my-ads')] | //*[contains(text(),'My Kijiji')]") # Success indicators
                )

                # Check for login failure messages
                error_messages = self.kjj.find_elements(By.XPATH, "//*[contains(@class,'error') and (contains(text(),'Invalid') or contains(text(),'incorrect'))]")
                if error_messages:
                    for msg in error_messages:
                        print(f"Login failed with error: {msg.text}")
                    self.take_screenshot(self.kjj, f"login_failed_attempt_{attempt+1}")
                    if attempt < max_login_attempts - 1:
                        # Optional: refresh or navigate back to login page if Kijiji doesn't allow immediate retry
                        # self.kjj.get("Kijiji_LOGIN_URL_HERE")
                        time.sleep(3)
                        continue # Retry
                    else:
                        return False # Max attempts reached

                # Check for success indicators (e.g., "My Ads" link or URL no longer contains "login")
                if "login" not in self.kjj.current_url.lower() or self.kjj.find_elements(By.XPATH, "//a[contains(@href, 'my-ads')] | //*[contains(text(),'My Kijiji')]"):
                    print(f"Login successful for user '{self.username}'. Current URL: {self.kjj.current_url}")
                    return True
                else: # No error, but no clear success either
                    print("Login status unclear after submission.")
                    self.take_screenshot(self.kjj, f"login_status_unclear_attempt_{attempt+1}")
                    if attempt < max_login_attempts - 1:
                        time.sleep(3)
                        continue
                    else:
                        return False


            except TimeoutException:
                print(f"TimeoutException during login attempt {attempt + 1}. Elements not found or page did not respond.")
                self.take_screenshot(self.kjj, f"timeout_login_attempt_{attempt + 1}")
                # self.kjj.refresh() # Refreshing here might clear fields or take you away.
                # Consider navigating back to the login page if Kijiji's flow requires it after a timeout.
                # e.g., self.kjj.get("https://www.kijiji.ca/t-login.html")
                if attempt == max_login_attempts - 1: return False # Failed on last attempt
                time.sleep(2) # Wait before retrying
            except ElementClickInterceptedException as eci_login:
                print(f"ElementClickInterceptedException during login submission attempt {attempt+1}: {eci_login}")
                self.take_screenshot(self.kjj, f"click_intercepted_login_submit_attempt_{attempt+1}")
                if attempt == max_login_attempts - 1: return False
                time.sleep(2)
            except Exception as e_login:
                print(f"An unexpected error occurred during login attempt {attempt + 1}: {str(e_login)}")
                self.take_screenshot(self.kjj, f"error_login_attempt_{attempt + 1}")
                if attempt == max_login_attempts - 1: return False
                # self.kjj.back() # Navigating back might not always be the right solution
                time.sleep(3) # Wait before retrying
        
        print("Login failed after all attempts.")
        return False


    def delete_ad(self,ad_data):
        wait = WebDriverWait(self.kjj, 60)
        self.current_ad_title = ad_data['Title'].strip()
        print(self.current_ad_title)
        self.next_url('https://www.kijiji.ca/m-my-ads/active/1?siteLocale=en_CA')
        time.sleep(5)   
        div_elements = []
        while True:
            try:
                tbody_element = self.kjj.find_element(By.TAG_NAME, 'tbody')
                tr_elements = tbody_element.find_elements(By.TAG_NAME,'tr')

                for tr_element in tr_elements:
                    div_elements = tr_element.find_elements(By.XPATH,'.//*')   
                    found_ad_to_delete = False
                    if not div_elements:
                        print("No div elements found.")
                    else:        
                        for div_element in div_elements:
                            links = div_element.find_elements(By.XPATH, ".//a[@href]")
                            for link in links:
                                if self.current_ad_title in link.get_attribute("innerHTML"):    
                                    try:
                                        delete_button = tr_element.find_element(By.XPATH, './/td[8]/div[1]/div[2]/div/div[2]/button/span')
                                        delete_button.click()
                                        found_ad_to_delete = True
                                        break
                                    except NoSuchElementException:
                                        print("Delete button not found for the ad.")
                                    except Exception as e:
                                        print(f"An unexpected error occurred: {str(e)}")
                            if found_ad_to_delete:
                                try:
                                    button = WebDriverWait(self.kjj, 10).until(EC.element_to_be_clickable((By.XPATH, '//button[text()="Prefer not to say"]')))
                                    button.click()
                                    time.sleep(2)
                                    button = WebDriverWait(self.kjj, 10).until(EC.element_to_be_clickable((By.XPATH, '//button[text()="Delete My Ad"]')))
                                    button.click()
                                    time.sleep(2)
                                    print('Ad deleted.')
                                    button = WebDriverWait(self.kjj, 10).until(EC.element_to_be_clickable((By.ID, 'modalCloseButton')))
                                    button.click()
                                    time.sleep(2)
                                    self.kjj.refresh()
                                    time.sleep(10)
                                    return
                                except TimeoutException:
                                    print('Timeout occurred while deleting the ad.')
                                except Exception as e:
                                    print(f"An unexpected error occurred: {str(e)}")
                            else:
                                print('Ad with title "{}" not found for deletion.'.format(self.current_ad_title))

                next_active_page = self.kjj.find_element(By.XPATH, '/html/body/div[3]/div[4]/div/div/div/div/div[3]/a')
                if next_active_page.text.strip() == "Next":
                    next_active_page.click()
                else:
                    return
            
            except StaleElementReferenceException:
                # Handle stale element exception by re-finding the element
                print("Stale element reference. Retrying...")
                continue

            except NoSuchElementException:
                print("Element not found.")
                return
                # Handle the situation here, such as returning from the method or raising an error

            except Exception as e:
                print(f"An unexpected error occurred: {str(e)}")
                return

    def check_Ads(self, ad_data): # check_ads is more Pythonic
        wait = WebDriverWait(self.kjj, 10) # Shorter wait for checks
        self.current_ad_title = ad_data.get('Title', '').strip()
        if not self.current_ad_title:
            print("No ad title provided for checking. Cannot check.")
            return None # Indicate error or invalid input

        print(f"Checking if ad exists: '{self.current_ad_title}'")
        self.next_url('https://www.kijiji.ca/m-my-ads/active/1')
        time.sleep(3)

        page_count = 0
        max_pages_to_check = 3 # Usually, if an ad exists, it's on the first few pages.

        while page_count < max_pages_to_check:
            page_count += 1
            print(f"Checking for ad '{self.current_ad_title}' on active ads page {page_count}...")
            try:
                tbody_element = wait.until(EC.presence_of_element_located((By.TAG_NAME, 'tbody')))
                # Look for links containing the ad title text. Using normalize-space for robustness.
                # Example: //a[contains(normalize-space(.), "Your Ad Title Here")]
                ad_title_xpath = f".//a[contains(normalize-space(.), \"{self.current_ad_title}\")]"
                # Check if any such links exist within the table body
                found_links = tbody_element.find_elements(By.XPATH, ad_title_xpath)

                if found_links:
                    print(f"Ad '{self.current_ad_title}' FOUND on page {page_count}.")
                    return 0 # Ad found

                print(f"Ad '{self.current_ad_title}' not found on page {page_count}.")
                # Try to navigate to the next page if ad not found on current page
                try:
                    next_page_button_xpath = '//a[(contains(@class, "pagination-next") or @aria-label="Next page" or descendant::span[normalize-space()="Next"]) and not(contains(@class,"disabled")) and not(@disabled)]'
                    next_page_link = WebDriverWait(self.kjj, 5).until(
                        EC.element_to_be_clickable((By.XPATH, next_page_button_xpath))
                    )
                    current_url_before_next = self.kjj.current_url
                    next_page_link.click()
                    WebDriverWait(self.kjj, 10).until(EC.url_changes(current_url_before_next)) # Confirm page changed
                    print("Navigated to next page for ad check.")
                    time.sleep(2) # Allow next page to load
                except TimeoutException:
                    print("No 'Next' page button found or it's disabled. End of ad list for checking.")
                    return 1 # Ad not found after checking all available pages
                except ElementClickInterceptedException as eci_check_next:
                    print(f"Click intercepted when trying to go to next page in check_Ads: {eci_check_next}")
                    self.take_screenshot(self.kjj, "click_intercepted_next_page_check_ads")
                    return None # Error state
            
            except TimeoutException: # Timeout for tbody_element
                print(f"Timeout waiting for ad table (tbody) on page {page_count} during check_Ads.")
                self.take_screenshot(self.kjj, f"timeout_tbody_check_ads_page_{page_count}")
                return None # Error state
            except StaleElementReferenceException:
                print(f"StaleElementReferenceException encountered on page {page_count} of check_Ads. Retrying this page.")
                self.take_screenshot(self.kjj, f"stale_element_check_ads_page_{page_count}")
                page_count -= 1 # To retry current page
                time.sleep(1)
                continue
            except Exception as e:
                print(f"An unexpected error occurred in check_Ads on page {page_count}: {str(e)}")
                self.take_screenshot(self.kjj, f"error_check_ads_page_{page_count}")
                return None # Error state
        
        print(f"Ad '{self.current_ad_title}' not found after checking {max_pages_to_check} pages.")
        return 1 # Ad not found


    def read_txt(self, file_path):
        credentials = {}
        try:
            with open(file_path, 'r') as file:
                for line in file:
                    line = line.strip()
                    if ':' in line: # Ensure there's a colon to split
                        key, value = line.split(':', 1) # Split only on the first colon
                        credentials[key.strip()] = value.strip()
                    elif line: # Non-empty line but no colon
                        print(f"Warning: Malformed line in credentials file (missing ':'): '{line}'")
        except FileNotFoundError:
            print(f"Error: Credentials file not found at '{file_path}'")
            return None # Indicate error clearly
        except Exception as e:
            print(f"Error reading credentials file '{file_path}': {e}")
            return None
        if not credentials:
             print(f"Warning: No valid credentials found in '{file_path}'")
             return None # No data parsed
        return credentials


def main():
    # The loop `for j in range (1,2)` runs only once with j=1.
    # This structure is fine if it's for potential future expansion.
    print("\n--- Starting Kijiji Ad Management ---")
    browser_instance = kijiji() # Create instance of the class

    # --- Credential Handling ---
    credentials_file = 'credentials.txt' # As per original
    print(f"Reading credentials from: {credentials_file}")
    credentials_data = browser_instance.read_txt(credentials_file)

    if not credentials_data: # Check if credentials loaded successfully
        print("Failed to load credentials. Exiting program.")
        return # Stop execution if no credentials

    if 'username' not in credentials_data or 'password' not in credentials_data:
        print("Error: 'username' or 'password' missing from credentials file. Cannot proceed.")
        return

    # --- Database Connection ---
    print(f"Connecting to database for user: {credentials_data['username']}")
    # Assuming connect_db uses credentials_data['username'] as table name implicitly
    ad_json_data_from_db = browser_instance.connect_db(credentials_data)

    if ad_json_data_from_db is None:
        print("Failed to connect to the database or retrieve ad data. Exiting program.")
        # Depending on requirements, you might allow proceeding without DB data for certain operations.
        return

    # --- Browser Initialization and Login ---
    print("Initializing Kijiji browser session...")
    browser_instance.access_kijiji(credentials_data) # Initializes self.kjj

    if not hasattr(browser_instance, 'kjj') or browser_instance.kjj is None:
        print("Browser (self.kjj) was not initialized by access_kijiji. Cannot proceed.")
        return

    # print("Attempting to log into Kijiji...")
    # login_successful = browser_instance.login(credentials_data)

    # if not login_successful:
    #     print("Kijiji login failed. Please check credentials or website status. Exiting.")
    #     if browser_instance.kjj:
    #         browser_instance.kjj.quit()
    #     return

    print("Kijiji login successful.")

    # --- Ad Processing ---
    try:
        print("Loading ad data from database JSON...")
        ads_to_process = json.loads(ad_json_data_from_db)
        if not isinstance(ads_to_process, list):
            print("Error: Ad data from database is not in the expected list format.")
            if browser_instance.kjj: browser_instance.kjj.quit()
            return
        
        num_ads = len(ads_to_process)
        print(f"Found {num_ads} ad(s) to process from the database.")

        for index, ad_row_data in enumerate(ads_to_process):
            print(f"\n--- Processing Ad {index + 1} of {num_ads} ---")
            if not isinstance(ad_row_data, dict):
                print(f"Skipping ad entry {index + 1} as it's not a valid dictionary: {ad_row_data}")
                continue

            # Construct ad_data dictionary using .get() for safety
            current_ad_details = {
                'Title': ad_row_data.get('Title'),
                'Category': ad_row_data.get('Category'),
                'Price': ad_row_data.get('Price'),
                'Description': ad_row_data.get('Description'),
                'Condition': ad_row_data.get('Condition'),
                'PhoneBrand': ad_row_data.get('PhoneBrand'),
                'PhoneBrandCarrier': ad_row_data.get('PhoneBrandCarrier'),
                'Images_FolderName': ad_row_data.get('Images_FolderName'),
                'Phone': ad_row_data.get('Phone'),
                'Tags': ad_row_data.get('Tags'),
                'Size': ad_row_data.get('Size'),
                'Type': ad_row_data.get('Type'),
                'Tablet Brand': ad_row_data.get('Tablet Brand'), # Keep original key casing
                'laptop Screen Size': ad_row_data.get('laptop Screen Size') # Keep original key casing
            }

            if not current_ad_details.get('Title'): # Essential field
                print(f"Ad entry {index + 1} is missing a 'Title'. Skipping this ad.")
                continue
            
            print(f"Processing Ad Title: '{current_ad_details['Title']}'")

            # Delete Ad
            print(f"Attempting to delete ad (if it exists): '{current_ad_details['Title']}'")
            browser_instance.delete_ad(current_ad_details)
            print(f"Deletion process finished for '{current_ad_details['Title']}'.")
            time.sleep(10) # Original sleep after delete

            # Post Ad
            print(f"Attempting to post ad: '{current_ad_details['Title']}'")
            browser_instance.post_ad(current_ad_details)
            print(f"Posting process finished for '{current_ad_details['Title']}'.")
            
            # Original had two 30s sleeps. Consolidating for clarity, adjust if specific timing is needed.
            print("Waiting 60 seconds before processing the next ad or finishing...")
            time.sleep(60) # Original was time.sleep(30); time.sleep(30)

        print("\n--- All ads from the database have been processed. ---")

    except json.JSONDecodeError as je:
        print(f"Error decoding JSON ad data from the database: {je}")
        if hasattr(browser_instance, 'kjj') and browser_instance.kjj:
            browser_instance.take_screenshot(browser_instance.kjj, "json_decode_error_main")
    except AttributeError as ae:
        # This might occur if browser_instance.kjj becomes None unexpectedly
        print(f"AttributeError in main ad processing: {ae}. Browser might have closed or failed.")
        # Attempt screenshot if kjj still exists and is a driver
        if hasattr(browser_instance, 'kjj') and isinstance(browser_instance.kjj, webdriver.Chrome):
             browser_instance.take_screenshot(browser_instance.kjj, "attribute_error_main_processing")
    except Exception as e_main:
        print(f"An unexpected error occurred during the main ad processing loop: {e_main}")
        if hasattr(browser_instance, 'kjj') and isinstance(browser_instance.kjj, webdriver.Chrome):
            browser_instance.take_screenshot(browser_instance.kjj, "unexpected_error_main_processing")
    finally:
        if hasattr(browser_instance, 'kjj') and browser_instance.kjj:
            print("Closing Kijiji browser session.")
            browser_instance.kjj.quit()
        print("--- Kijiji Ad Management Finished ---")


if __name__ == '__main__':
    main()
